# Selenium imports
import glob
from selenium import webdriver
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.common.exceptions import NoSuchElementException, TimeoutException, InvalidSessionIdException
# Converting language to the "sound" (ie Chinese > Chinese Pinyin)
from pypinyin import lazy_pinyin, Style
import pykakasi
from hangul_romanize import Transliter
from hangul_romanize.rule import academic
import pythainlp
# other tools
import re
import sqlite3
import sys
# Output to EXCEL
import os
import shutil
from openpyxl import load_workbook
import datetime
import time
# PyTube instead of YouTube API
from pytube import Search as ytSearch
from pytube import YouTube

transliter = Transliter(academic)
cid_list = []


def db_init(chosen_year):
    con = sqlite3.connect('kp_albums.db')
    cur = con.cursor()
    try:
        cur.execute(f'DROP TABLE IF EXISTS y{chosen_year}')
    except sqlite3.OperationalError:
        pass
    try:
        cur.execute('DROP TABLE IF EXISTS mv')
        cur.execute('CREATE TABLE mv(video, vname, song, artist, channel, cid, verified)')
    except sqlite3.OperationalError:
        pass
    cur.execute(f"CREATE TABLE y{chosen_year}(artist, album, songs, url, mv, mvverified)")
    con.commit()
    return con, cur


def main(kpop_year):
    # Open kpopping
    driver.get(f"https://kpopping.com/musicalbums/year-{kpop_year}")

    # on webpage, find all artists and albums in the list
    artists = driver.find_elements(By.CSS_SELECTOR, 'a[href*="/profiles/idol"]')
    albums = driver.find_elements(By.CSS_SELECTOR, 'a[href*="/musicalbum"]')
    artist_group = driver.find_elements(By.CSS_SELECTOR, 'a[href*="/profiles/group"]')
    # create artists and albums lists
    artists_list = []
    albums_list_raw = []

    # append lists with text from previous elements
    for artist in artists:
        artists_list.append(artist.text)
    for artist in artist_group:
        artists_list.append(artist.text)
    for album in albums:
        albums_list_raw.append(album.text)
        # edit list to match the formatting of kpopping links
        albums_list1 = [f"{kpop_year}-" + item.replace(" ", "-") for item in albums_list_raw]
        # does NOT remove - (hyphens)
        albums_list2 = [translate_symbol(item) for item in albums_list1]
        albums_list = [remove_special_characters(item) for item in albums_list2]

        # error can be ignored
    merged_list = list(zip(artists_list, albums_list))
    # convert to dictionary to make sorting easier later
    merged_dict = {}
    for artist, album in merged_list:
        if artist not in merged_dict:
            merged_dict[artist] = [album]
        else:
            merged_dict[artist].append(album)
    attempted_urls = []
    it_num = 1
    for artist, album in merged_dict.items():
        print(artist)
        try:
            re_init = 0
            kpopping_songs(album, attempted_urls, it_num, re_init, artist)
        except InvalidSessionIdException:
            re_init = 1
            kpopping_songs(album, attempted_urls, it_num, re_init, artist)


def kpopping_songs(album, attempted_urls, it_num, re_init, artist):
    if re_init == 1:
        print('FireFox has crashed. Re-initializing..')
        time.sleep(3)
        redriver = webdriver.Firefox(options=options, service=service)
        time.sleep(1)
        for f in album:
            if f'https://kpopping.com/musicalbum/{f}'.lower() in attempted_urls:
                it_num += 1
                f = f'https://kpopping.com/musicalbum/{f}{it_num}'
            try:
                redriver.get(f'https://kpopping.com/musicalbum/{f}')
                attempted_urls.append(f'https://kpopping.com/musicalbum/{f}'.lower())
                print(f'https://kpopping.com/musicalbum/{f}')
            except TimeoutException:
                attempt = 0
                while True:
                    attempt += 1
                    time.sleep(30)
                    try:
                        redriver.get(f'https://kpopping.com/musicalbum/{f}')
                        break
                    except TimeoutException:
                        if attempt == 6:
                            print('TIMED OUT. 6 ATTEMPTS MADE OVER 3 MINUTES.')
                            redriver.quit()
                            quit()
            try:
                redriver.find_element(By.CSS_SELECTOR, '.fa-compact-disc')
                songs = grab_songs()
                # DATABASE IS IN FORMAT artist, album, songs, url, friendly_name
                # db_album used to format album name into a more readable format
                db_album = str(album[4:]).replace('-', '')
                for item in songs:
                    # FNAME
                    print(item)
                    cur.execute(f"INSERT INTO y{chosen_year_q} (artist,album,songs) VALUES (?,?,?)",
                                (artist, db_album, item))
                    kp_db.commit()
            except NoSuchElementException:
                no_album = 'N/A'
                no_song = 'N/A'
                cur.execute(f"INSERT INTO y{chosen_year_q} (artist,album,songs) VALUES (?,?,?)", (artist, no_album,
                                                                                                  no_song))
                kp_db.commit()
    else:
        for f in album:
            if f'https://kpopping.com/musicalbum/{f}'.lower() in attempted_urls:
                it_num += 1
                f = f'https://kpopping.com/musicalbum/{f}{it_num}'
            try:
                driver.get(f'https://kpopping.com/musicalbum/{f}')
                attempted_urls.append(f'https://kpopping.com/musicalbum/{f}'.lower())
                print(f'https://kpopping.com/musicalbum/{f}')
            except TimeoutException:
                attempt = 0
                while True:
                    attempt += 1
                    time.sleep(30)
                    try:
                        driver.get(f'https://kpopping.com/musicalbum/{f}')
                        break
                    except TimeoutException:
                        if attempt == 6:
                            print('TIMED OUT. 6 ATTEMPTS MADE OVER 3 MINUTES.')
                            driver.quit()
                            quit()
            try:
                driver.find_element(By.CSS_SELECTOR, '.fa-compact-disc')
                songs = grab_songs()
                # DATABASE IS IN FORMAT artist, album, songs, url, friendly_name
                # db_album used to format album name into a more readable format
                db_album = str(album[4:]).replace('-', '')
                for item in songs:
                    # FNAME
                    print(item)
                    cur.execute(f"INSERT INTO y{chosen_year_q} (artist,album,songs) VALUES (?,?,?)",
                                (artist, db_album, item))
                    kp_db.commit()
            except NoSuchElementException:
                no_album = 'N/A'
                no_song = 'N/A'
                cur.execute(f"INSERT INTO y{chosen_year_q} (artist,album,songs) VALUES (?,?,?)", (artist, no_album,
                                                                                                  no_song))
                kp_db.commit()


def grab_songs():
    songs_section = driver.find_element(By.CSS_SELECTOR, '.tracks')
    songs_section_list = songs_section.find_elements(By.CLASS_NAME, 'title-wr')
    songs_raw = []
    for song in songs_section_list:
        songs_raw.append(song.text)
    return songs_raw


def remove_special_characters(s):
    # Replace spaces with hyphens - this is to try to get around the issue with ' in album names
    # had to add more as some were coming out with -- and --- in the names, this has fixed all issues of this kind
    # SO FAR
    pattern = r'[\u2665\(),!]'
    s = re.sub(pattern, '', s)
    replace_list_space = ["'"]
    replace_list_dash = [' ', '.', '_', '+', '&', 'ː', '=', '÷', '/', '#', '?', '...', '’', '…', '---', '--']
    replace_list_empty = [':', '[', ']', '"']
    s = s.replace(':)', '2')
    s = s.replace('Ø', 'O')
    for i in replace_list_space:
        s = s.replace(i, ' ')
    for i in replace_list_empty:
        s = s.replace(i, '')
    for i in replace_list_dash:
        s = s.replace(i, '-')
    # Removes - if at the end of string
    s = re.sub(r'-\Z', '', s)
    s = s.replace('Ⅱ', 'II')
    s = s.replace('Ⅲ', 'III')
    s = s.replace('ᐸ', '<')
    s = s.replace('ᐳ', '>')
    s = s.replace('<', '%E1%90%B8')
    s = s.replace('>', '%E1%90%B3')
    # Exception due to inconsistency from kpopping
    if s == ('https://kpopping.com/musicalbum/2023-The-Seasons-VolII-10-%E1%90%B8AKMU-s-Long-Day-Long-'
             'Night%E1%90%B3-Re-Wake-x-Chuu'):
        s = s.replace('%E1%90%B8', '')
        s = s.replace('%E1%90%B3', '')
    return s


def translate_symbol(s):
    chinese_chars = re.findall(r'[\u4e00-\u9fff]+', s)
    for chinese_char in chinese_chars:
        pinyin_char = ''.join(lazy_pinyin(chinese_char, style=Style.NORMAL))
        s = s.replace(chinese_char, pinyin_char)
        s = s.replace('[', '-')
        s = s.replace(']', '')
    japanese_chars = re.findall(r'[\u3040-\u30FF\u31F0-\u31FF\u3400-\u4DBF\u4E00-\u9FFF\uF900-\uFAFF]+', s)
    for japanese_char in japanese_chars:
        kks = pykakasi.kakasi()
        result = kks.convert(japanese_char)
        for item in result:
            romaji_char = ''.join(item['hepburn'])
            s = s.replace(japanese_char, romaji_char)
        s = s.replace('[', '-')
        s = s.replace(']', '')
    korean_chars = re.findall(r'[\uac00-\ud7af\u1100-\u1112\u1161-\u1175\u11A8-\u11C2]+', s)
    for korean_char in korean_chars:
        romanized_char = ''.join(transliter.translit(korean_char))
        s = s.replace(korean_char, romanized_char)
        s = s.replace('[', '-')
        s = s.replace(']', '')
    thai_chars = re.findall(r'[\u0e00-\u0e7f]+', s)
    for thai_char in thai_chars:
        thairom_char = ''.join(pythainlp.romanize(thai_char))
        s = s.replace(thai_char, thairom_char)
        s = s.replace('[', '-')
        s = s.replace(']', '')
    return s


def music_video():
    mv_searches = cur.execute(f"SELECT artist, songs FROM y{chosen_year_q}").fetchall()
    iteration = 0
    batch_execute = 1
    for mv_res in mv_searches:
        print(batch_execute)
        if batch_execute == 1:
            jf = open(f'response_files/format_response{iteration}.txt', 'w+', encoding='utf-8')
            artist = mv_res[0]
            song = mv_res[1]
            print(f'{song} {artist}')
            search_term = f'{mv_res} {artist}'
            search_result = ytSearch(search_term).results
            x = 0
            for i in search_result:
                i = str(i)
                if x < 11:
                    start_index = i.find("videoId=") + len("videoId=")
                    end_index = i.find(">", start_index)
                    video_id = i[start_index:end_index]
                    search_id = f'https://www.youtube.com/watch?v={video_id}'
                    youtube_return = YouTube(search_id)
                    cid_line = youtube_return.channel_id
                    channel_user = youtube_return.channel_url
                    start_index = channel_user.find("@") + len("@")
                    channel_name = channel_user[start_index:]
                    upload_date = youtube_return.publish_date.year
                    vname_line = youtube_return.title
                    print(video_id)
                    jf.write(f'video_id:{video_id}\nchannel_id:{cid_line}\nvideo_name:{vname_line}\nchannel_'
                             f'name:{channel_name}\nupload_year:{upload_date}\n[debug]search_term:{search_term}\n'
                             f'\n')
                    x += 1
        else:
            artist = mv_res[0]
            song = mv_res[1]
            print(f'{song} {artist}')
            search_term = f'{mv_res} {artist}'
            search_result = ytSearch(search_term).results
            x = 0
            for i in search_result:
                i = str(i)
                if x < 11:
                    start_index = i.find("videoId=") + len("videoId=")
                    end_index = i.find(">", start_index)
                    video_id = i[start_index:end_index]
                    search_id = f'https://www.youtube.com/watch?v={video_id}'
                    youtube_return = YouTube(search_id)
                    cid_line = youtube_return.channel_id
                    channel_user = youtube_return.channel_url
                    start_index = channel_user.find("@") + len("@")
                    channel_name = channel_user[start_index:]
                    upload_date = youtube_return.publish_date.year
                    vname_line = youtube_return.title
                    print(video_id)
                    jf.write(f'video_id:{video_id}\nchannel_id:{cid_line}\nvideo_name:{vname_line}\nchannel_'
                             f'name:{channel_name}\nupload_year:{upload_date}\n[debug]search_term:{search_term}\n'
                             f'\n')
                    x += 1
        batch_execute += 1
        if batch_execute == 10:
            batch_execute = 1
            jf.close()
            mv_database(filename=f'response_files/format_response{iteration}.txt', song=song, artist=artist)
            iteration += 1


def mv_database(filename, song, artist):
    with open(filename, 'r', encoding='utf-8') as input_file:
        lines = input_file.readlines()
    vid_lines = [line.split(':')[1].strip() for line in lines if 'video_id' in line]
    cid_lines = [line.split(':')[1].strip() for line in lines if 'channel_id' in line]
    vname_lines = [line.split(':')[1].strip() for line in lines if 'video_name' in line]
    cname_lines = [line.split(':')[1].strip() for line in lines if 'channel_name' in line]
    year_check = [line.split(':')[1].strip() for line in lines if 'upload_year' in line]
    for vid, cid, vname, cname, eyear in zip(vid_lines, cid_lines, vname_lines, cname_lines, year_check):
        mv_exist = cur.execute("""SELECT video FROM mv WHERE video=?""", (vid,)).fetchone()
        if artist not in vname:
            print(f'Video title does not contain {artist}. Including video but may be inaccurate. Title: {vname}')
            verified = 3
        elif song not in vname:
            print(f'Video title does not contain {song}. Including video but may be inaccurate. Title: {vname}')
            verified = 3
        else:
            verified = 0
        cid_list.append(cid)
        if mv_exist:
            if eyear[:4] != chosen_year_q:
                print(f'Video uploaded in wrong year: {eyear}. Skipping this video..')
                continue
            else:
                print(f'Video with ID {vid} already exists in database. Skipping {song} by {artist}')
                video_type(vname, vid)
        else:
            print(f'[DEBUG] EXTRACTED YEAR = {eyear}\nCHOSEN YEAR = {chosen_year_q}')
            if eyear[:4] == chosen_year_q:
                cur.execute("""INSERT INTO mv (video, vname, song, artist, channel, cid, verified)
                VALUES (?, ?, ?, ?, ?, ?, ?)""", (vid, vname, song, artist, cname, cid, verified,))
                print(f'Video with ID {vid} and name {vname} added to database')
                video_type(vname, vid)
            else:
                print(f'Video uploaded in wrong year: {eyear}. Skipping this video..')
                continue
    kp_db.commit()


def video_type(vid, yt_id):
    teaser_words = ['teaser', 'teaser video', 'spoiler', 'highlight medley']
    reaction_words = ['reaction']
    mv_words = ['mv', 'music video', 'm/v', '뮤비', 'official video']
    dance_words = ['dance', 'practice', '춤', '관행', '춤 연습', '나 춤 연습해']
    perf_words = ['performance', 'stage', 'live', 'inkigayo', "[it's Live]", '[BE ORIGINAL]', '1theKILLPO', 'tour']
    lyric_words = ['lyric', 'color coded', 'lyrics', 'line distribution', '가사']
    youtube_link = f'https://www.youtube.com/watch?v={yt_id}'
    if check_string_for_phrases(vid, teaser_words):
        cur.execute("""INSERT INTO video_type (name, vid_id, vid_type) VALUES (?, ?, ?)""",
                    (vid, youtube_link, 'teaser'))
    elif check_string_for_phrases(vid, reaction_words):
        cur.execute("""INSERT INTO video_type (name, vid_id, vid_type) VALUES (?, ?, ?)""",
                    (vid, youtube_link, 'reaction'))
    elif check_string_for_phrases(vid, mv_words):
        cur.execute("""INSERT INTO video_type (name, vid_id, vid_type) VALUES (?, ?, ?)""",
                    (vid, youtube_link, 'mv'))
    elif check_string_for_phrases(vid, dance_words):
        cur.execute("""INSERT INTO video_type (name, vid_id, vid_type) VALUES (?, ?, ?)""",
                    (vid, youtube_link, 'dance'))
    elif check_string_for_phrases(vid, perf_words):
        cur.execute("""INSERT INTO video_type (name, vid_id, vid_type) VALUES (?, ?, ?)""",
                    (vid, youtube_link, 'performance'))
    elif check_string_for_phrases(vid, lyric_words):
        cur.execute("""INSERT INTO video_type (name, vid_id, vid_type) VALUES (?, ?, ?)""",
                    (vid, youtube_link, 'lyric'))
    else:
        cur.execute("""INSERT INTO video_type (name, vid_id, vid_type) VALUES (?, ?, ?)""",
                    (vid, youtube_link, 'other'))


def check_string_for_phrases(input_string, target_phrases):
    for phrase in target_phrases:
        if re.search(r'\b' + re.escape(phrase) + r'\b', input_string, flags=re.IGNORECASE):
            return True
    return False


def channel_verified():
    run_number = 0
    for i in cid_list:
        is_verified = cur.execute("""SELECT verified FROM mv WHERE cid=?""", (i,)).fetchone()
        if is_verified == 'tbc':
            run_number += 1
            format_url = f'https://www.youtube.com/channel/{i}'
            if verify_element_exist(format_url, run_number):
                print(f'{format_url} IS VERIFIED')
                cur.execute("""UPDATE mv SET verified = 1 WHERE cid=?""", (i,))
            else:
                print(f'{format_url} IS NOT VERIFIED')
                cur.execute("""UPDATE mv SET verified = 0 WHERE cid=?""", (i,))
        else:
            continue
    kp_db.commit()
    for i in cid_list:
        # Check if manual verification is required
        ver_value = cur.execute("""SELECT verified FROM mv WHERE cid=?""", (i,)).fetchone()
        if ver_value == 2:
            manual_verify(i)
        else:
            pass


def verify_element_exist(url, run_number):
    # CSS Selector
    verified_badge = ('ytd-channel-name.ytd-c4-tabbed-header-renderer > ytd-badge-supported-renderer:nth-child(2) '
                      '> div:nth-child(1) > yt-icon:nth-child(1) > yt-icon-shape:nth-child(1) > icon-shape:nth-child(1)'
                      ' > div:nth-child(1)')
    # CSS Selector
    artist_badge = ('ytd-channel-name.ytd-c4-tabbed-header-renderer > ytd-badge-supported-renderer:nth-child(2) '
                    '> div:nth-child(1) > yt-icon:nth-child(1) > yt-icon-shape:nth-child(1) > icon-shape:nth-child(1) '
                    '> div:nth-child(1) > svg:nth-child(1)')
    # SELENIUM LOAD YOUTUBE
    # check if elements exist (to see if cookies page or not)
    # cookie_accept appears if you open the YouTube main page. keeping in code for now in case it crops up again
    cookie_accept = ('ytd-button-renderer.ytd-consent-bump-v2-lightbox:nth-child(2) > yt-button-shape:nth-child(1) '
                     '> button:nth-child(1) > yt-touch-feedback-shape:nth-child(2) > div:nth-child(1) > '
                     'div:nth-child(2)')
    cookie_accept_2 = ('.KZ9vpc > form:nth-child(3) > div:nth-child(1) > div:nth-child(1) > '
                       'button:nth-child(1) > span:nth-child(4)')
    driver.get(url)
    if run_number == 1:
        try:
            WebDriverWait(driver, 10).until(ec.presence_of_element_located((By.CSS_SELECTOR, cookie_accept_2)))
            driver.find_element(By.CSS_SELECTOR, cookie_accept_2).click()
        except NoSuchElementException:
            try:
                driver.find_element(By.CSS_SELECTOR, 'input.ytd-searchbox').click()
            except NoSuchElementException:
                print('Failed to load page. Exiting..')
                driver.quit()
                sys.exit()
    else:
        pass
    if driver.find_elements(By.CSS_SELECTOR, artist_badge) or driver.find_elements(By.CSS_SELECTOR, verified_badge):
        print('VERIFIED')
        return True
    else:
        time.sleep(5)
        if driver.find_elements(By.CSS_SELECTOR, artist_badge) or driver.find_elements(By.CSS_SELECTOR, verified_badge):
            print('VERIFIED')
            return True
        else:
            print('NOT VERIFIED')
            return False


def manual_verify(cid):
    do_manual_verify = ('The script has complete. There are some videos uploaded by channels without a \n'
                        'verification tick. Would you like to iterate through the list to manually verify these \n'
                        'channels? NOTE: This is used to help pick out a channel that would upload the correct \n'
                        'video and not a fan made video. (Y/N): ')
    if do_manual_verify.casefold() == 'y':
        print('Please answer questions with either a Y for yes or an N for no: ')
        for i in cid:
            channel_name = cur.execute("""SELECT cname FROM mv WHERE cid=?""", (i,)).fetchone()
            yes_channel = input(f'Should {channel_name} (www.youtube.com/channel/{cid}) be verified (Y/N)?:')
            if yes_channel.casefold() == 'y':
                cur.execute("""UPDATE mv SET verified=1 WHERE cid=?""", (cid,))
            elif yes_channel.casefold() == 'n':
                cur.execute("""UPDATE mv SET verified=0 WHERE cid=?""", (cid,))
            else:
                print('Input invalid. You will need to manually change this in the database or re-run script.')
                cur.execute("""UPDATE mv SET verified=2 WHERE cid=?""", (cid,))
    elif do_manual_verify.casefold() == 'n':
        print('No problem, if you change your mind you can manually open the "kp_albums.db" database and open the\n'
              '"mv" table and change the value in the "verified" column to a 1 for verified, 0 for not. Or run '
              '"alter_verification.bat" to use a script to do this.')
        time.sleep(1)
        print('Cleaning up..')
        time.sleep(2)
        print('Exiting script..')
        time.sleep(1)
        driver.quit()
        sys.exit()
    else:
        print('Input not recognised. Expected result either: Y or N\nPlease try again..\n\n')
        manual_verify(cid)


def output_excel():
    if not os.path.isdir('output'):
        os.mkdir('output')
    template_path = 'output_template.xlsx'
    workbook_name = f'{datetime.date.today().strftime("%y%m%d")}_{datetime.datetime.now().strftime("%H%M%S")}.xlsx'
    output_path = os.path.join('output', workbook_name)
    shutil.copy(template_path, output_path)
    workbook = load_workbook(output_path)
    worksheet = workbook.active
    worksheet.title = f'KPOP_{chosen_year_q}'
    row = 2
    col = 1
    artists = cur.execute("""SELECT artist FROM mv""").fetchall()
    already_read = set()
    for artist in artists:
        if artist not in already_read:
            already_read.add(artist)
            artist_list = list(artist)
            artist = str(artist_list[0])
            songs = cur.execute(f"""SELECT songs FROM y{chosen_year_q} WHERE artist=?""", (artist,)).fetchall()
            already_read_song = set()
            for ii in songs:
                if ii not in already_read_song:
                    song_list = list(ii)
                    sel_song = str(song_list[0])
                    worksheet.cell(row=row, column=col).value = artist
                    worksheet.cell(row=row, column=col + 1).value = sel_song
                    already_read_song.add(ii)
                    video_res = cur.execute(f"""SELECT video FROM mv WHERE song=?""", (ii)).fetchall()
                    video_list = list(video_res)
                    mv_list = []
                    dance_list = []
                    performance_list = []
                    teaser_list = []
                    reaction_list = []
                    lyric_list = []
                    other_list = []
                    for vid_sql in video_list:
                        pattern = r'[\'(),]'
                        vid = re.sub(pattern, '', str(vid_sql))
                        vid_type = cur.execute("""SELECT vid_type FROM video_type WHERE vid_id=?""", (
                            f'https://www.youtube.com/watch?v={vid}',)).fetchone()
                        pattern = r'[\'(),]'
                        vid_type = re.sub(pattern, '', str(vid_type))
                        if vid_type == 'mv':
                            mv_list.append(f'https://www.youtube.com/watch?v={vid}, ')
                        elif vid_type == 'dance':
                            dance_list.append(f'https://www.youtube.com/watch?v={vid}, ')
                        elif vid_type == 'performance':
                            performance_list.append(f'https://www.youtube.com/watch?v={vid}, ')
                        elif vid_type == 'teaser':
                            teaser_list.append(f'https://www.youtube.com/watch?v={vid}, ')
                        elif vid_type == 'reaction':
                            reaction_list.append(f'https://www.youtube.com/watch?v={vid}, ')
                        elif vid_type == 'lyric':
                            lyric_list.append(f'https://www.youtube.com/watch?v={vid}, ')
                        elif vid_type == 'other':
                            other_list.append(f'https://www.youtube.com/watch?v={vid}, ')
                        else:
                            print(f'N/A - {artist} potentially has no videos found. Please check output. [FAULTY'
                                  f' - CAN LIkELY BE IGNORED AT THIS TIME]')
                    type_list = [mv_list, dance_list, performance_list, teaser_list, reaction_list, lyric_list,
                                 other_list]
                    col_num = 4
                    for i in type_list:
                        format_list = ''.join(i)
                        worksheet.cell(row=row, column=col + col_num).value = format_list
                        col_num += 1
                    verified_sql = cur.execute("""SELECT verified FROM mv WHERE video=?""", (vid,)).fetchone()
                    if verified_sql:
                        verified = verified_sql[0]
                        worksheet.cell(row=row, column=col + 10).value = verified
                    else:
                        worksheet.cell(row=row, column=col + 10).value = "Not Verified"
                    row += 1
    workbook.save(output_path)


if __name__ == '__main__':
    import timeit
    start = timeit.default_timer()
    print('Cleaning files..')
    clean_files = glob.glob('response_files/*')
    for i in clean_files:
        os.remove(i)
    chosen_year_q = input('Year to get data for (format: YYYY e.g 2024): ')
    print('Loading database..')
    kp_db, cur = db_init(chosen_year_q)
    kp_db.commit()
    try:
        cur.execute(f"CREATE TABLE mv(video, vname, song, artist, channel, cid, verified)")
        kp_db.commit()
    except sqlite3.OperationalError:
        pass
    cur.execute("DROP TABLE IF EXISTS video_type")
    kp_db.commit()
    cur.execute(f"CREATE TABLE video_type(name, vid_id, vid_type)")
    kp_db.commit()
    # Setup geckodriver (Firefox)
    print('Launching Firefox via Geckodriver in HEADLESS mode..')
    options = Options()
    options.add_argument("-headless")
    geckodriver_path = 'geckodriver.exe'
    service = Service(executable_path=geckodriver_path)
    driver = webdriver.Firefox(options=options, service=service)
    print(f'Checking kpopping list for year {chosen_year_q}..')
    main(chosen_year_q)
    print('List created. Checking music videos..')
    music_video()
    print('Data grab complete. Checking channel verification status.. (May take a while.. '
          'Approx. 6-7 seconds per video)')
    channel_verified()
    driver.quit()
    output_excel()
    stop = timeit.default_timer()
    print(f'Total runtime: {stop - start}')
