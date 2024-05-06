import shutil
from openpyxl import load_workbook


# noinspection PyShadowingNames
# Above is to stop pycharm complaining about names shadowing from outer scope, this is due to the: if __name__ section
def output_excel(os, datetime, chosen_year_q, cur, re):
    if not os.path.isdir('output'):
        os.mkdir('output')
    template_path = 'output_template.xlsx'
    workbook_name = f'{datetime.date.today().strftime("%y%m%d")}_{datetime.datetime.now().strftime("%H%M%S")}.xlsx'
    output_path = os.path.join('output', workbook_name)
    shutil.copy(template_path, output_path)
    workbook = load_workbook(output_path)
    worksheet = workbook.active
    worksheet.title = f'KPOP_{chosen_year_q}'
    row = 2
    row_ver = 2
    col = 1
    artists = cur.execute("""SELECT artist FROM mv""").fetchall()
    already_read = set()
    for artist in artists:
        if artist not in already_read:
            already_read.add(artist)
            artist_list = list(artist)
            artist = str(artist_list[0])
            songs = cur.execute(f"""SELECT songs FROM y{chosen_year_q} WHERE artist=?""", (artist,)).fetchall()
            already_read_song = set()
            for ii in songs:
                if ii not in already_read_song:
                    song_list = list(ii)
                    sel_song = str(song_list[0])
                    worksheet.cell(row=row, column=col).value = artist
                    worksheet.cell(row=row, column=col + 1).value = sel_song
                    already_read_song.add(ii)
                    video_res = cur.execute(f"""SELECT video FROM mv WHERE song=?""", (ii)).fetchall()
                    video_list = list(video_res)
                    mv_list = []
                    dance_list = []
                    performance_list = []
                    teaser_list = []
                    reaction_list = []
                    lyric_list = []
                    other_list = []
                    verify_list = []
                    for vid_sql in video_list:
                        pattern = r'[\'(),]'
                        vid = re.sub(pattern, '', str(vid_sql))
                        vid_type = cur.execute("""SELECT vid_type FROM video_type WHERE vid_id=?""", (
                            f'https://www.youtube.com/watch?v={vid}',)).fetchone()
                        pattern = r'[\'(),]'
                        vid_type = re.sub(pattern, '', str(vid_type))
                        if vid_type == 'mv':
                            mv_list.append(f'https://www.youtube.com/watch?v={vid}, ')
                        elif vid_type == 'dance':
                            dance_list.append(f'https://www.youtube.com/watch?v={vid}, ')
                        elif vid_type == 'performance':
                            performance_list.append(f'https://www.youtube.com/watch?v={vid}, ')
                        elif vid_type == 'teaser':
                            teaser_list.append(f'https://www.youtube.com/watch?v={vid}, ')
                        elif vid_type == 'reaction':
                            reaction_list.append(f'https://www.youtube.com/watch?v={vid}, ')
                        elif vid_type == 'lyric':
                            lyric_list.append(f'https://www.youtube.com/watch?v={vid}, ')
                        elif vid_type == 'other':
                            other_list.append(f'https://www.youtube.com/watch?v={vid}, ')
                        else:
                            print(f'N/A - {artist} potentially has no videos found. Please check output. [FAULTY'
                                  f' - CAN LIkELY BE IGNORED AT THIS TIME]')
                        check_if_verified(vid, verify_list, cur)

                    type_list = [mv_list, dance_list, performance_list, teaser_list, reaction_list, lyric_list,
                                 other_list]
                    col_num = 4
                    for i in type_list:
                        format_list = ''.join(i)
                        worksheet.cell(row=row, column=col + col_num).value = format_list
                        col_num += 1
                    row += 1
                    for i in verify_list:
                        if i == 1:
                            worksheet.cell(row=row_ver, column=col + 10).value = "Verified"
                        elif i == 2:
                            worksheet.cell(row=row_ver, column=col + 10).value = "Manual review"
                        elif i == 3:
                            worksheet.cell(row=row_ver, column=col + 10).value = 3
                        else:
                            worksheet.cell(row=row_ver, column=col + 10).value = "Verification error"
                        row_ver += 1
    workbook.save(output_path)


def check_if_verified(vid, verify_list, cur):
    verified_sql = cur.execute("""SELECT verified FROM mv WHERE video=?""", (vid,)).fetchone()
    verify_list.append(verified_sql[0])


if __name__ == '__main__':
    import os
    import datetime
    import re
    from db_init import db_connect
    chosen_year_q = input('Year to get data for (format: YYYY e.g 2024): ')
    cur = db_connect(chosen_year_q, is_main=0)
    output_excel(os, datetime, chosen_year_q, cur, re)
